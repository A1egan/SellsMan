import openpyxl, json, re

SRC = r"C:/Users/小田/Desktop/强化群学员跟进.xlsx"

BLUE   = 'BACEFD'          # 浅蓝 -> 99体验
PURPLE = 'ECE2FE'         # 浅紫 -> 算法营1
PINK_FILL = 'FBBFBC'      # 粉填充 -> 算法营2
PINK_FONT = 'F01D94'      # 粉字体 -> 算法营2
GRAY_L = 'DEE0E3'         # 浅灰备注 -> 低意向
GRAY_D = '8F959E'         # 深灰备注 -> 低意向
GREEN  = '8EE085'         # 绿色 -> 直接删除（不导入）

TAGS = {
    '99体验':  'tag_99',
    '算法营1': 'tag_algo1',
    '算法营2': 'tag_algo2',
}

def norm(rgb):
    if not rgb: return None
    rgb = str(rgb)
    if rgb.startswith('FF'): rgb = rgb[2:]
    return rgb.upper()

def campaign(id_fill, id_font):
    f = norm(id_fill); ft = norm(id_font)
    if f == BLUE:   return '99体验'
    if f == PURPLE: return '算法营1'
    if f == PINK_FILL or ft == PINK_FONT: return '算法营2'
    return None

wb = openpyxl.load_workbook(SRC)
users = {}   # id -> user dict
stats = {'99体验':0,'算法营1':0,'算法营2':0,'lowinterest':0,'pending_untagged':0,'other_color_untagged':0}
other_colors = {}

for ws in wb.worksheets:
    maxc = ws.max_column; maxr = ws.max_row
    for r in range(3, maxr+1):
        for c in range(1, maxc, 2):   # odd col = ID
            idv = ws.cell(r,c).value
            if idv is None: continue
            sid = str(idv).strip()
            if not sid: continue
            note = ws.cell(r, c+1).value
            note = '' if note is None else str(note).strip()
            idf = ws.cell(r,c).fill
            id_fill = idf.fgColor.rgb if (idf and idf.fgColor and idf.fgColor.rgb) else None
            idft = ws.cell(r,c).font
            id_font = idft.color.rgb if (idft and idft.color and idft.color.rgb and isinstance(idft.color.rgb,str)) else None
            nf = ws.cell(r, c+1).fill
            note_fill = nf.fgColor.rgb if (nf and nf.fgColor and nf.fgColor.rgb) else None

            # 绿色底色 -> 直接删除（不导入）
            if norm(id_fill) == GREEN:
                continue

            tag = campaign(id_fill, id_font)
            low = norm(note_fill) in (GRAY_L, GRAY_D)
            col = 'lowinterest' if low else 'pending'

            # other (non-rule) colors track
            f = norm(id_fill)
            if f not in (BLUE, PURPLE, PINK_FILL, None) and not low:
                other_colors.setdefault(f, 0)
                other_colors[f] += 1

            if sid in users:
                u = users[sid]
                if not u['note'] and note: u['note'] = note
                if tag and tag not in u['tagNames']: u['tagNames'].append(tag)
                if col == 'lowinterest': u['column'] = 'lowinterest'
                continue

            u = {
                'id': 'u_'+sid,
                'number': sid,
                'column': col,
                'note': note,
                'replied': False,
                'tags': [],
                'tagNames': [tag] if tag else [],
                'createdAt': 0,
                'updatedAt': 0,
            }
            users[sid] = u

# finalize tags + stats
out = []
for sid, u in users.items():
    tags = [TAGS[t] for t in u['tagNames'] if t in TAGS]
    u['tags'] = tags
    out.append(u)
    # stats
    if u['column'] == 'lowinterest':
        stats['lowinterest'] += 1
    elif u['tags']:
        tn = [k for k,v in TAGS.items() if v==u['tags'][0]][0]
        stats[tn] += 1
    else:
        stats['pending_untagged'] += 1
    del u['tagNames']

print("Total unique students:", len(out))
print("Stats:", json.dumps(stats, ensure_ascii=False))
print("Other (non-rule) ID fill colors seen (count, not yet tagged):")
for k,v in sorted(other_colors.items(), key=lambda x:-x[1]):
    print("   ", k, v)

# write seed json
seed = {
    'tags': [
        {'id':'tag_99','name':'99体验','color':'#3b82f6'},   # 蓝色
        {'id':'tag_algo1','name':'算法营1','color':'#a78bfa'}, # 浅紫色
        {'id':'tag_algo2','name':'算法营2','color':'#6d28d9'}, # 深紫色
    ],
    'users': out,
}
with open(r"C:/Users/小田/WorkBuddy/2026-08-20-10-35-27/seed_import.json","w",encoding="utf-8") as f:
    json.dump(seed, f, ensure_ascii=False)
print("Wrote seed_import.json")
